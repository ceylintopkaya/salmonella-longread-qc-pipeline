# -*- coding: utf-8 -*-
from Bio import SeqIO
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

input_file = "data/salmonella.fastq"
results = []

for record in SeqIO.parse(input_file, "fastq"):
    length = len(record.seq)
    g = record.seq.count("G")
    c = record.seq.count("C")
    gc_content = (g + c) / length * 100
    quality_scores = record.letter_annotations["phred_quality"]
    mean_quality = sum(quality_scores) / len(quality_scores)
    results.append({
        "read_id": record.id,
        "length": length,
        "gc_content": gc_content,
        "mean_quality": mean_quality
    })

df = pd.DataFrame(results)
df.to_csv("results/read_stats.csv", index=False)

print("Total reads: " + str(len(df)))
print("Mean read length: " + str(round(df['length'].mean(), 1)))
print("Mean GC content: " + str(round(df['gc_content'].mean(), 1)))
print("Mean quality score: " + str(round(df['mean_quality'].mean(), 1)))
print("Raw data saved to results/read_stats.csv!")

summary = pd.DataFrame({
    "Metric": ["Read Length (bp)", "GC Content (%)", "Quality Score"],
    "Mean": [round(df["length"].mean(),1), round(df["gc_content"].mean(),1), round(df["mean_quality"].mean(),1)],
    "Median": [round(df["length"].median(),1), round(df["gc_content"].median(),1), round(df["mean_quality"].median(),1)],
    "Min": [round(df["length"].min(),1), round(df["gc_content"].min(),1), round(df["mean_quality"].min(),1)],
    "Max": [round(df["length"].max(),1), round(df["gc_content"].max(),1), round(df["mean_quality"].max(),1)],
    "Std Dev": [round(df["length"].std(),1), round(df["gc_content"].std(),1), round(df["mean_quality"].std(),1)]
})
summary.to_csv("results/summary_stats.csv", index=False)
print("Summary statistics saved!")

fig, ax = plt.subplots(figsize=(10, 2))
ax.axis('off')
table = ax.table(cellText=summary.values, colLabels=summary.columns, cellLoc='center', loc='center')
table.auto_set_font_size(False)
table.set_fontsize(11)
table.scale(1.2, 1.8)
plt.title("Salmonella - Summary Statistics", fontsize=13, fontweight='bold', pad=20)
plt.savefig("results/summary_table.png", dpi=150, bbox_inches='tight')
print("PNG table saved!")

df_clean = df.copy()
df_clean["gc_content"] = df_clean["gc_content"].round(2)
df_clean["mean_quality"] = df_clean["mean_quality"].round(2)
df_clean.columns = ["Read ID", "Length (bp)", "GC Content (%)", "Quality Score"]

with pd.ExcelWriter("results/report.xlsx", engine="openpyxl") as writer:
    summary.to_excel(writer, sheet_name="Summary Statistics", index=False)
    df_clean.to_excel(writer, sheet_name="Raw Data", index=False)

wb = load_workbook("results/report.xlsx")
ws = wb["Raw Data"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
for col in ws.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4
wb.save("results/report.xlsx")
print("Excel report saved!")
print("\nAll files created successfully!")
